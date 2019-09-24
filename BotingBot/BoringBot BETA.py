import openpyxl
import os

excelFile = input('Database > ')
wb = openpyxl.load_workbook(excelFile)
sheet = wb['Foglio1']
ws = wb.active

database = {}



#---------------------------------------------------------------

def IfExistsRename():
	if os.path.exists(before) == True:
		after = database[before]
		os.rename(before, after)
		print(f'{before} -> {after}')

def CheckFolders():
	num = 100
	while num < 107:
		folder = "CANON" + str(num)
		os.chdir(os.path.join(rootFolder, folder))
		#print(os.path.join(rootFolder, folder))
		IfExistsRename()
		num += 1
	

#---------------------------------------------------------------
#					ACQUISIZIONE DATABASE
#---------------------------------------------------------------

for i in range(1,11):
	photoNum1 = sheet.cell(row=i, column=2).value
	photoNum2 = sheet.cell(row=i, column=3).value
	
	name = sheet.cell(row=i, column=1).value.title()
	
	name = name.split()
	lname = len(name)

	if lname < 3:
		nameForm1 = str(name[0]) + "_" + str(name[1]) + "_" +"1.png"
		nameForm2 = str(name[0]) + "_" + str(name[1]) + "_" +"2.png"
	elif lname==3:
		nameForm1 = str(name[0]) + "_" + str(name[1]) + "_" + str(name[-1]) + "_" + "1.png"
		nameForm2 = str(name[0]) + "_" + str(name[1]) + "_" + str(name[-1]) + "_" + "2.png"
	else:
		break
	
	photoNum1 = str(photoNum1) + '.png'
	photoNum2 = str(photoNum2) + '.png'
	
	
	if not photoNum1 in database:
		database[photoNum1] = nameForm1

	if not photoNum2 in database:
		database[photoNum2] = nameForm2
	


#---------------------------------------------------------------
#---------------------------------------------------------------


print(''' 
RenameBot
By SamueleMaker 2019
''')

c = 1
for elemento in database:
	print(f'{c}) {elemento} => {database[elemento]}')
	c += 1
print(' ')
rootFolder = input('Foto > ')

for before in database:
	
	CheckFolders()
	
print(' ')
input('Lavoro completato, premi un tasto per uscire. ')





