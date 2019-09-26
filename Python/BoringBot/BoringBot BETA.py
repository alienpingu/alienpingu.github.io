import openpyxl
import os

print(''' 
RenameBot
By SamueleMaker 2019
''')


excelFile = input('Database > ')
wb = openpyxl.load_workbook(excelFile)
sheet = wb['Foglio1']
ws = wb.active

database = {}

datamove = {}


#---------------------------------------------------------------

def CheckFolders():
	f = ["2005" ,"2007", "2008", "2009-2010", "2011", "2012", "2013-2014", "Squadra femminile", "Juniores", "Prima squadra", "Staff"]
	n = 0
	while n < 11:
		folder = f[n]
		os.chdir(os.path.join(rootInput, folder))
		print(os.path.join(rootInput, folder))
		IfExistsRename()
		n += 1

def IfExistsRename():
	if os.path.exists(before) == True:
		after = database[before]
		os.rename(before, after)
		print(f'{before} -> {after}')




#---------------------------------------------------------------
#                   ACQUISIZIONE DATABASE
#---------------------------------------------------------------

for i in range(1,188):
	#cartella = sheet.cell(row=i, column=1).value
	titolo = sheet.cell(row=i, column=3).value.title()
	name = sheet.cell(row=i, column=2).value.title()
	photoNum1 = sheet.cell(row=i, column=4).value
	photoNum2 = sheet.cell(row=i, column=5).value
	
	

	if not titolo == "-":
		titolo = titolo.replace(" ", "_")
		nameForm1 = name.replace(" ", "_") + "_" + titolo + "_1.JPG"
		nameForm2 = name.replace(" ", "_") + "_" + titolo + "_2.JPG"
	
	else:
		titolo = titolo.replace(" ", "_")
		nameForm1 = str(name.replace(" ", "_")) + "_1.JPG"
		nameForm2 = str(name.replace(" ", "_")) + "_2.JPG"
	




	photoNum1= "IMG_" + str(photoNum1) + '.JPG'
	photoNum2= "IMG_" + str(photoNum2) + '.JPG'
	
	
	if not photoNum1 in database:
		database[photoNum1] = nameForm1
		
	if not photoNum2 in database:
		database[photoNum2] = nameForm2
		
	



#---------------------------------------------------------------
#---------------------------------------------------------------



c = 1
for elemento in database:
	print(f'{c}) {elemento} => {database[elemento]} ')
	c += 1


print(' ')
rootInput = input('Foto > ')
os.chdir(rootInput)
#Rinominazione

for before in database:
	CheckFolders()	



#End Program

print(' ')
input('Lavoro completato, premi un tasto per uscire. ')





