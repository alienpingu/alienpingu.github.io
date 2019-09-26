import os
import openpyxl

path = 'C:/Users/samue/Documents/GitHub/alienpingu.github.io/BoringBot/INPUT/'
moveto = "C:/Users/samue/Documents/GitHub/alienpingu.github.io/BoringBot/OUTPUT/"

excelFile = input('Database > ')
wb = openpyxl.load_workbook(excelFile)
sheet = wb['Foglio1']
ws = wb.active

dataMove = {}

for i in range(1,10):
	
	name = sheet.cell(row=i, column=1).value
	folder = sheet.cell(row=i, column=2).value


	if not name in dataMove:
		dataMove[name] = folder


for n in dataMove:
	if os.path.exists(n) == True:
		src = path + n
		dst = path + dataMove[folder] + n


files = os.listdir(path)
files.sort()

print(files)
for f in files:

	src = path + f
	dst = moveto + folder + f
	os.rename(src,dst)




input()
